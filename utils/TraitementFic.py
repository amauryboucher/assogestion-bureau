import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal
from loguru import logger

from utils.database_operation import insert_adherent_to_database, import_operation_to_database, \
    rapprochement_adherent_operation
from utils.import_adherent_controls import global_control


def none_to_str(value):
    if value:
        return value
    else:
        return ""


class TraiterFichier(QThread):
    def __init__(self, item, Session, type_import, parent=None):
        try:
            QThread.__init__(self, parent)
            self.item = item
            self.Session = Session
            self.type = type_import
            self.dic_error = {
                "ADH_CTRL_001": "Le titre de l’adhérent ne peut pas être vide",
                "ADH_CTRL_002": "Le nom et le prénom de l’adhérent sont des données obligatoire",
                "ADH_CTRL_003": "L’adresse de l’adhérent est obligatoire",
                "ADH_CTRL_004": "Le code postal de l’adhérent ne doit pas être vide",
                "ADH_CTRL_005": "La ville de l’adhérent ne doit pas être vide",
                "ADH_CTRL_006": "Le format de l’email n’est pas correct",
                "ADH_CTRL_008": "La cotisation indiquée n'existe pas dans la base de données",
                "ADH_ENR_009": "Impossible d'enrichir automatiquement la cotisation de l'adhérent",
                "ORG_CTRL_001": "Le titre de l’organisation ne peut pas être vide",
                "ORG_CTRL_002": "Le nom et le prénom de l'organisation sont des données obligatoire",
                "ORG_CTRL_003": "L’adresse de l’organisation est obligatoire",
                "ORG_CTRL_004": "Le code postal de l’organisation ne doit pas être vide",
                "ORG_CTRL_005": "La ville de l’organisation ne doit pas être vide",
                "ORG_CTRL_006": "Le format de l’email n’est pas correct",
                "ORG_CTRL_008": "La cotisation indiquée n'existe pas dans la base de données",
                "ORG_ENR_009": "Impossible d'enrichir automatiquement la cotisation de l'organisation",
                "": ""
            }
        except Exception as err:
            logger.error(err)

    line_count_changed = pyqtSignal(int)
    line_count_max = pyqtSignal(int)
    step_changed = pyqtSignal(str)
    flag_fin = pyqtSignal(list)

    def run(self):
        try:
            liste_error = []
            logger.info("Lancement traitement")
            logger.info(self.item)

            wb = openpyxl.load_workbook(filename=self.item[0])
            ws = wb.worksheets[0]
            self.line_count_max.emit(ws.max_row)
            for i in range(2, ws.max_row + 1):
                ligne = (
                    none_to_str(ws[f"A{i}"].value),
                    none_to_str(ws[f"B{i}"].value),
                    none_to_str(ws[f"C{i}"].value),
                    none_to_str(ws[f"D{i}"].value),
                    none_to_str(ws[f"E{i}"].value),
                    none_to_str(ws[f"F{i}"].value),
                    none_to_str(ws[f"G{i}"].value),
                    none_to_str(ws[f"H{i}"].value),
                    none_to_str(ws[f"I{i}"].value),
                    none_to_str(ws[f"J{i}"].value)
                )
                self.step_changed.emit("Contrôle des données à intégrer pour le fichier adhérents")
                res = global_control(i, ligne)
                if res != "":
                    logger.error(f"- Ligne {i} --> {res}: {self.dic_error.get(res)}")
                    liste_error.append(f"Ligne {i} --> {res}: {self.dic_error.get(res)}\n")
                else:
                    res = insert_adherent_to_database(self.Session, i, ligne)
                    if res != "":
                        logger.info(res)
                        liste_error.append(f"Ligne {i} --> {res}\n")
                self.line_count_changed.emit(i)
            ws = wb.worksheets[1]
            self.line_count_max.emit(ws.max_row)
            self.line_count_changed.emit(0)
            for j in range(2, ws.max_row + 1):
                ligne = (
                    none_to_str(ws[f"A{j}"].value),
                    none_to_str(ws[f"B{j}"].value),
                    none_to_str(ws[f"C{j}"].value),
                )
                res = import_operation_to_database(self.Session, ligne)
                if res != "":
                    pass
                self.line_count_changed.emit(j)
            wb.close()
            self.line_count_max.emit(1)
            self.line_count_changed.emit(0)
            rapprochement_adherent_operation(self.Session)
            self.line_count_changed.emit(1)
            #send_import_report_adh(getpass.getuser(), liste_error)
            self.flag_fin.emit([True, 0, None])
        except Exception as err:
            logger.error(err)