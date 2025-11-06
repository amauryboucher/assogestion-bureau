import os
from datetime import datetime

import openpyxl
import win32com
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QTableWidgetItem
from sqlalchemy import update
from win32com.universal import com_error

from IHM import facture
from utils.facture_controls import *
from utils.format_input import format_date_to_database
from utils.message import pop_up
from utils.models import assogestion_client, assogestion_facture, assogestion_ligne_facture, \
    assogestion_counter_facture


class DetailFactAPPV2(QtWidgets.QMainWindow, facture.Ui_facture):
    finished = pyqtSignal(bool)

    def __init__(self, Session, edit, fac, parent=None):
        super(DetailFactAPPV2, self).__init__(parent)
        try:

            self.save = False
            self.Session = Session
            self.edit = edit
            self.icon = QtGui.QIcon()
            self.icon.addPixmap(QtGui.QPixmap("logo_abo_tech.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
            self.setWindowIcon(self.icon)
            self.setupUi(self)
            self.showMaximized()
            self.application = None
            self.button_select_client.setIcon(QIcon('rechercher.png'))
            self.button_select_client.clicked.connect(self.init_client)
            self.button_save_client.clicked.connect(self.sauver_client)
            self.button_del_fac.clicked.connect(self.close)
            self.client_cp.textChanged.connect(lambda: ENR_CLI_003(self))
            self.button_save_fac.clicked.connect(self.sauver_facture)
            self.tab_ligne_facture.itemChanged.connect(self.save_changes)
            self.actionAjouter_une_ligne.triggered.connect(self.add_ligne)
            self.actionSupprimer_une_ligne.triggered.connect(self.supp_ligne)
            self.actionGenerer_Facture.triggered.connect(self.export_invoice_to_pdf)
            if not self.edit:
                self.data_dict = {}
                self.data_ligne_dict = {}
                self.generate_invoice()
            else:
                self.data_dict = fac
                self.init_app()

            self.dic_error = {
                "CF_CLI_001": "Le code client ne peut pas être vide",
                "CF_CLI_002": "La raison sociale du client ne peut pas être vide",
                "CF_CLI_004": "L'adresse du client ne peut pas être vide",
                "CF_CLI_005": "Le code postal du client ne peut pas être vide",
                "CF_CLI_006": "La ville du client ne peut pas être vide",
                "CF_CLI_007": "Le format de l'adresse email du client est incorrect",
                "CM_FAC_008": "Le total d'une ligne de facture doit être supérieur à 0",
                "CF_FAC_009": "la facture doit être rattachée à un client",
                "": ""
            }


        except Exception as err:
            logger.error(err)


    def generate_invoice(self):
        s = self.Session()
        inv_line = 1
        try:
            counter = s.query(assogestion_counter_facture).first()
            new_invoice_num = int(counter.last_invoice_num) + 1
            logger.info(f"Dernier numéro: FC0{counter.last_invoice_num}, Nouveau numéro: FC0{new_invoice_num}")
            self.data_dict = {
                "inv_no": new_invoice_num,
                "inv_number_full": f"FC0{new_invoice_num}",
                "code_client": "CLXXXX",
                "date": datetime.now(),
                "date_echeance" : datetime.now()
            }
            new_invoice = assogestion_facture(**self.data_dict)
            s.add(new_invoice)
            counter.last_invoice_num += 1
            s.commit()
            self.data_ligne_dict = {
                'inv_no' : self.data_dict.get('inv_no'),
                'inv_line_number': inv_line,
                'code' : "code",
                'desc' : "description",
                'qty' : 0,
                'prix_unitaire' : 0.00,
            }
            new_invoice_ligne = assogestion_ligne_facture(**self.data_ligne_dict)
            s.add(new_invoice_ligne)
            s.commit()
            self.init_app()
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))
        finally:
            s.close()

    def annuler_saisie(self):
        s = self.Session()
        try:
            counter = s.query(assogestion_counter_facture).first()
            del_ligne_fac = s.query(assogestion_ligne_facture).filter_by(inv_no=self.data_dict.get('inv_no')).delete()
            s.commit()
            del_fac = s.query(assogestion_facture).filter_by(inv_no=self.data_dict.get('inv_no')).delete()
            if counter.last_invoice_num > 0:
                counter.last_invoice_num -= 1
            s.commit()
        except Exception as err:
            logger.error(err)
        s.close()
    def init_app(self):
        s = self.Session()
        invoice = s.query(assogestion_facture).filter_by(inv_no=self.data_dict.get('inv_no')).first()
        s.close()
        try:
            self.num_facture_edit.setText(str(invoice.inv_number_full))
            self.date_facture_edit.setDate(invoice.date)
            self.date_echeance_facture.setDate(invoice.date_echeance)
            self.code_client_edit.setText(invoice.code_client)
            self.add_liv_edit.setPlainText(invoice.add_liv)
            self.order_edit.setText(invoice.order_number)
            if invoice.order_date:
                self.order_date_edit.setDate(invoice.order_date)
            if self.code_client_edit.text() not in ("", "CLXXXX"):
                self.init_client()
            self.init_ligne_fac()
            self.init_pied_fac()
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))

    def init_ligne_fac(self):
        s = self.Session()
        try:
            inv_no = int(self.num_facture_edit.text().split("FC0")[1])
            res = s.query(assogestion_ligne_facture).filter_by(inv_no=inv_no).order_by(assogestion_ligne_facture.inv_line_number).all()
            self.tab_ligne_facture.clear()
            self.tab_ligne_facture.setRowCount(0)
            self.tab_ligne_facture.setColumnCount(6)
            self.tab_ligne_facture.setHorizontalHeaderLabels(["Numéro", "Code", "Description", "Quantité", "Prix unitaire", "Total"])
            if res:
                for i, r in enumerate(res):
                    self.tab_ligne_facture.insertRow(i)
                    result = r.inv_line_number, r.code, r.desc, r.qty, r.prix_unitaire, float(r.qty) * r.prix_unitaire
                    for j, res in enumerate(result):
                        item = QTableWidgetItem(f'{result[j]}')
                        self.tab_ligne_facture.setItem(i, j, item)
                        if j == 0 or j == 5:
                            item.setFlags(QtCore.Qt.ItemIsEditable)

            self.tab_ligne_facture.resizeColumnsToContents()
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))
        s.close()
    def init_pied_fac(self):
        s = self.Session()
        inv_no = int(self.num_facture_edit.text().split("FC0")[1])
        res = s.query(assogestion_facture).filter_by(inv_no=inv_no).first()
        res_ligne = s.query(assogestion_ligne_facture).filter_by(inv_no=inv_no).order_by(
            assogestion_ligne_facture.inv_line_number).all()
        total_ht = 0
        total_ttc = 0
        total_a_payer = 0
        if res:
            if res_ligne:
                for r in res_ligne:
                    total_ht += float(r.qty) * float(r.prix_unitaire)
                total_ttc = total_ht
                total_a_payer = total_ht
            else:
                total_ht = res.total_ht
                total_ttc = res.total_ttc
                total_a_payer = res.total_a_payer
        self.tab_pied_facture.setItem(0, 3,QTableWidgetItem(f'{total_ht}'))
        self.tab_pied_facture.setItem(0, 5, QTableWidgetItem(f'{total_ttc}'))
        self.tab_pied_facture.setItem(0, 6, QTableWidgetItem(f'{total_a_payer}'))
        s.close()
    def init_client(self):
        s = self.Session()
        try:
            if self.code_client_edit.text() not in("CLXXXX", ""):
                client = s.query(assogestion_client).filter_by(id_client=self.code_client_edit.text()).first()
                if client:
                    self.client_rs.setText(client.raison_sociale)
                    self.client_adresse.setText(client.adresse_client)
                    self.client_adresse_comp.setText(client.adresse_comp_client)
                    self.client_cp.setText(client.cp_client)
                    self.client_mail.setText(client.mail_client)
                else:
                    self.label_error_client.setText("Ce code client n'existe pas dans la base de données. Veuillez saisir ses infos et cliquer sur Sauver")
            elif self.code_client_edit.text() == "":
                self.vider_client()
                self.label_error_client.setText("")
            else:
                self.label_error_client.setText("Impossible de rechercher par ce code client")
                self.vider_client()
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))
        finally:
            s.close()

    def vider_client(self):
        self.client_rs.setText("")
        self.label_client_rs.setText("")
        self.client_adresse.setText("")
        self.label_add_error.setText("")
        self.client_adresse_comp.setText("")
        self.label_cp.setText("")
        self.client_cp.setText("")
        self.label_ville.setText("")
        self.cb_ville.clear()
        self.label_mail.setText("")
        self.client_mail.setText("")

    def sauver_client(self):
        liste_error = []
        s = self.Session()
        nb_error = 0
        try:
            err_001 = CF_CLI_001(self)
            if err_001 != "":
                nb_error+=1
                liste_error.append(f"{err_001}: {self.dic_error.get(err_001)}")
            err_002 = CF_CLI_002(self)
            if err_002 != "":
                nb_error +=1
                liste_error.append(f"{err_002}: {self.dic_error.get(err_002)}")
            err_004 = CF_CLI_004(self)
            if err_004 != "":
                nb_error += 1
                liste_error.append(f"{err_004}: {self.dic_error.get(err_004)}")
            err_005 = CF_CLI_005(self)
            if err_005 != "":
                nb_error += 1
                liste_error.append(f"{err_005}: {self.dic_error.get(err_005)}")
            err_006 = CF_CLI_006(self)
            if err_006 != "":
                nb_error +=1
                liste_error.append(f"{err_006}: {self.dic_error.get(err_006)}")
            err_007 = CF_CLI_007(self)
            if err_007 != "":
                nb_error +=1
                liste_error.append(f"{err_007}: {self.dic_error.get(err_007)}")
            self.label_error_client.setText(self.dic_error.get(err_001))
            self.label_client_rs.setText(self.dic_error.get(err_002))
            self.label_add_error.setText(self.dic_error.get(err_004))
            self.label_cp.setText(self.dic_error.get(err_005))
            self.label_ville.setText(self.dic_error.get(err_006))
            self.label_mail.setText(self.dic_error.get(err_007))
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))

        if nb_error == 0:
            try:
                data_client = {
                    "id_client": self.code_client_edit.text(),
                    "raison_sociale": self.client_rs.text(),
                    "adresse_client": self.client_adresse.text(),
                    "adresse_comp_client": self.client_adresse_comp.text(),
                    "cp_client": self.client_cp.text(),
                    "ville_client": self.cb_ville.currentText(),
                    "mail_client": self.client_mail.text()
                }
                client = s.query(assogestion_client).filter_by(id_client=self.code_client_edit.text()).first()
                if client:
                    s.execute(update(assogestion_client).filter_by(id_client=client.id_client).values(**data_client))
                    s.commit()
                else:
                    new_client = assogestion_client(**data_client)
                    s.add(new_client)
                    s.commit()
                inv_dic = {
                    "code_client": self.code_client_edit.text()
                }
                s.execute(update(assogestion_facture).filter_by(inv_number_full=self.num_facture_edit.text()).values(**inv_dic))
                s.commit()
                self.success_facture("Enregistrement effectué", "Le client a bien été sauvegardé")
            except Exception as err:
                logger.error(err)
                self.error_facture("Une erreur est survenue", str(err))
        else:
            message = ""
            for i in liste_error:
                message += f"{i}\n"
            self.error_facture("Enregistrement impossible", f"Il reste {nb_error} erreur(s) dans le formulaire\n{message}")


    def add_ligne(self):
        s = self.Session()
        try:
            inv_no = int(self.num_facture_edit.text().split("FC0")[1])
            inv_line = s.query(assogestion_ligne_facture).filter_by(inv_no=inv_no).order_by(assogestion_ligne_facture.inv_line_number.desc()).first().inv_line_number+1

            data_ligne_dict = {
                'inv_no': self.data_dict.get('inv_no'),
                'inv_line_number': inv_line,
                'code': "code",
                'desc': "description",
                'qty': 0,
                'prix_unitaire': 0.00,
            }
            new_invoice_ligne = assogestion_ligne_facture(**data_ligne_dict)
            s.add(new_invoice_ligne)
            s.commit()
            self.init_ligne_fac()
            self.init_pied_fac()
        except Exception as err:
            logger.error(err)
            self.error_facture("add_ligne:Une erreur est survenue", str(err))

    def supp_ligne(self):
        s = self.Session()
        try:
            inv_line_number = self.tab_ligne_facture.item(self.tab_ligne_facture.selectedItems()[0].row(), 0).text()
            inv_no = int(self.num_facture_edit.text().split("FC0")[1])
            del_number = s.query(assogestion_ligne_facture).filter_by(inv_no=inv_no, inv_line_number=inv_line_number).delete()
            s.commit()
            logger.info(del_number)
            if del_number > 0:
                self.init_ligne_fac()
                self.init_pied_fac()
                self.success_facture("Suppression de ligne réussie", f"La ligne {inv_line_number} de la facture {self.num_facture_edit.text()} a été supprimée")
        except Exception as err:
            logger.error(err)
            self.error_facture("supp_ligne: Une erreur est survenue", str(err))
        finally:
            s.close()
    def sauver_facture(self):
        s = self.Session()
        nb_error = 0
        liste_error = []
        err_008 = CM_FAC_008(self)
        if err_008:
            self.label_montant_error.setText(self.dic_error.get(err_008))
            nb_error += 1
            liste_error.append(self.dic_error.get(err_008))
        err_009 = CF_FAC_009(self)
        if err_009:
            self.label_error_client.setText(self.dic_error.get(err_009))
            nb_error += 1
            liste_error.append(self.dic_error.get(err_009))
        if nb_error > 0:
            self.error_facture("sauver_facture : Enregistrement impossible", f"Il reste {nb_error} erreur(s) dans le formulaire")
        else:
            try:
                inv_no = int(self.num_facture_edit.text().split("FC0")[1])
                date_facture = format_date_to_database(self.date_facture_edit.text())
                date_echeance = format_date_to_database(self.date_facture_edit.text())
                total_ht = float(self.tab_pied_facture.item(0, 3).text())
                total_ttc = float(self.tab_pied_facture.item(0, 5).text())
                total_a_payer = float(self.tab_pied_facture.item(0, 6).text())
                order_number = self.order_edit.text()
                order_date = format_date_to_database(self.order_date_edit.text())
                add_liv = self.add_liv_edit.toPlainText()
                data_dict = {
                    "inv_no": inv_no,
                    "inv_number_full": self.num_facture_edit.text(),
                    "code_client": self.code_client_edit.text(),
                    "date": date_facture,
                    "date_echeance": date_echeance,
                    "total_ht": total_ht,
                    "total_ttc": total_ttc,
                    "total_a_payer": total_a_payer,
                    "order_number": order_number,
                    "order_date": order_date,
                    "add_liv": add_liv
                }
                s.execute(update(assogestion_facture).filter_by(inv_no=inv_no).values(**data_dict))
                s.commit()
                self.success_facture("Enregistrement réussi", f"La facture n°{data_dict.get('inv_number_full')} a été sauvegardée")
                self.save = True
                self.close()
            except Exception as err:
                s.close()
                logger.error(err)
                self.error_facture("sauver_facture : Une erreur est survenue", str(err))
    def error_facture(self, message, texte):
        pop_up(
            self,
            "AssoGestion - AMSE - Détail Facture",
            message,
            texte,
            "ERROR"
        )
    def success_facture(self, message, texte):
        pop_up(
            self,
            "AssoGestion - AMSE - Détail Facture",
            message,
            texte,
            "INFO"
        )
    def closeEvent(self, a0):
        if not self.edit and not self.save:
            self.annuler_saisie()
        self.finished.emit(True)
        a0.accept()

    def save_changes(self):
        self.tab_ligne_facture.resizeColumnsToContents()
        inv_line_number = None
        code = None
        desc = None
        qty = None
        prix_unitaire = None
        data_ligne = {
            'inv_line_number' : "",
            'code' : "",
            'desc' : "",
            'qty' : "",
            'prix_unitaire' : ""
        }
        s = self.Session()
        try:
            total_ht = 0.00
            qty = 0
            prix_unitaire = 0.00

            for currentQTableWidgetItem in self.tab_ligne_facture.selectedItems():
                inv_line_number = self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 0).text()
                code = self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 1).text()
                desc = self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 2).text()
                qty = self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 3).text()
                prix_unitaire = self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 4).text()
                self.tab_ligne_facture.item(currentQTableWidgetItem.row(), 5).setText(str(float(qty) * float(prix_unitaire)))
            self.init_pied_fac()
            data_ligne['inv_line_number'] = inv_line_number
            data_ligne['code'] = code
            data_ligne['desc'] = desc
            data_ligne['qty'] = qty
            data_ligne['prix_unitaire'] = prix_unitaire
            inv_no = int(self.num_facture_edit.text().split("FC0")[1])
            ligne = s.query(assogestion_ligne_facture).filter_by(inv_no=inv_no, inv_line_number=inv_line_number).first()
            if ligne:
                s.execute(update(assogestion_ligne_facture).filter_by(inv_line_number=inv_line_number, inv_no=inv_no).values(**data_ligne))
                s.commit()
        except Exception as err:
            logger.error(err)
            self.error_facture("save_changes :Une erreur est survenue", str(err))
        finally:
            s.close()

    def export_invoice_to_pdf(self):
        s = self.Session()
        try:
            fac = s.query(assogestion_facture).filter_by(inv_number_full=self.num_facture_edit.text()).first()
            if fac:
                client = s.query(assogestion_client).filter_by(id_client=fac.code_client).first()
                if client:
                    wb = openpyxl.load_workbook(f"{os.getcwd()}\\modele\\modele_facture.xlsx")
                    ws = wb.active
                    ws["F3"].value = fac.inv_number_full
                    ws["G3"].value = datetime.strftime(fac.date, "%d-%m-%Y")
                    ws["H3"].value = fac.code_client
                    ws["F9"].value = client.raison_sociale
                    ws["F10"].value = client.adresse_client
                    ws["F11"].value = client.adresse_comp_client
                    ws["F12"].value = f"{client.cp_client} {client.ville_client}"
                    ws["F14"].value = fac.add_liv
                    ws["D17"].value = fac.order_number
                    ws["D18"].value = fac.order_date
                    lignes_facture = s.query(assogestion_ligne_facture).filter_by(inv_no=int(self.num_facture_edit.text().split("FC0")[1])).all()
                    start = 21
                    for i, l in enumerate(lignes_facture):
                        ws[f"B{start+i}"].value = l.code
                        ws[f"C{start+i}"].value = l.qty
                        ws[f"D{start + i}"].value = l.desc
                        ws[f"G{start + i}"].value = l.prix_unitaire
                        ws[f"H{start + i}"].value = l.qty * l.prix_unitaire
                    ws["C37"].value = fac.total_ht
                    ws["E37"].value = fac.total_ht
                    ws["G37"].value = fac.total_ttc
                    ws["H37"].value = fac.total_a_payer
                    wb.save(f"{os.getcwd()}\\Export\\Facture\\{fac.inv_number_full}.xlsx")
                    wb.close()
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    try:
                        logger.info('Start conversion to PDF')
                        # Open
                        wb = excel.Workbooks.Open(f"{os.getcwd()}\\Export\\Facture\\{fac.inv_number_full}.xlsx")
                        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
                        ws_index_list = [1]
                        wb.WorkSheets(ws_index_list).Select()
                        # Save
                        wb.ActiveSheet.ExportAsFixedFormat(0, f"{os.getcwd()}\\Export\\Facture\\{fac.inv_number_full}.pdf")
                    except com_error as e:
                        logger.error('failed.')
                        self.error_facture("Une erreur est survenue", str(e))
                    else:
                        logger.info('Succeeded.')
                    finally:
                        wb.Close()
                        excel.Quit()
                    self.success_facture("Export de la facture terminé", "La facture a été exportée en pdf")
                else:
                    self.error_facture("Une erreur est survenue","Impossible de retrouver le client de la facture")
            else:
                self.error_facture("Une erreur est survenue","Impossible d'exporter la facture")
        except Exception as err:
            logger.error(err)
            self.error_facture("Une erreur est survenue", str(err))
        finally:
            s.close()